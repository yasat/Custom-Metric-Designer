from django.contrib.auth.decorators import user_passes_test
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponseRedirect, HttpResponse
from django.urls import reverse
from core.models import *
from .AI_model import *

from django.utils.html import strip_tags

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from collections import defaultdict

FAIRNESS_METRIC_MAP = {
    "demographic_parity": demographic_parity,
    "equal_opportunity": equal_opportunity,
    "predictive_equality": predictive_equality,
    "outcome_test": outcome_test,
    "equalized_odds": equalized_odds,
    "conditional_statistical_parity": conditional_statistical_parity,
    "counterfactual_fairness": counterfactual_fairness,
}

def render_metric_preview(metric, staging):
    if isinstance(metric, DesignExistingMetrics):
        return render_existing_preview(metric)

    elif isinstance(metric, DesignCombineMetrics):
        return render_combine_existing_preview(staging)

    elif isinstance(metric, DesignCustomOwnMetric):
        return render_custom_preview(staging)

    elif isinstance(metric, DesignProceduralMetric):
        return render_procedural_preview(staging)

    elif isinstance(metric, AffordabilityDesign):
        return render_affordability_preview(metric)

    return "Preview not available"

def render_existing_preview(metric):
    protected = []
    non_protected = []

    try:
        protected_df, non_protected_df, prot_cols, nonprot_cols = parse_feature_tags(
            metric.features,
            return_columns=True
        )
    except Exception as e:
        return f"<div class='alert alert-danger'>Error parsing features: {str(e)}</div>", "Unfair"

    threshold = (metric.threshold or 0.0)
    metric_id = (metric.metric or "").lower().strip()

    try:
        if metric_id == "counterfactual_fairness":
            if not (prot_cols and nonprot_cols):
                raise ValueError("Counterfactual fairness requires 1 protected and 1 non-protected feature.")
            col_a, col_b = prot_cols[0], nonprot_cols[0]
            assessment = counterfactual_fairness(data, col_a, col_b, threshold=threshold)

        elif metric_id == "conditional_statistical_parity":
            fn = FAIRNESS_METRIC_MAP[metric_id]
            assessment = fn(protected_df, non_protected_df, threshold, group_cols=[])
        else:
            fn = FAIRNESS_METRIC_MAP.get(metric_id, FAIRNESS_METRIC_MAP["demographic_parity"])
            assessment = fn(protected_df, non_protected_df, threshold)

    except Exception as e:
        return f"<div class='alert alert-danger'>Metric computation error: {str(e)}</div>", "Unfair"

    for item in metric.features.split(","):
        item = item.strip()
        if not item:
            continue
        parts = item.split("[")
        label = parts[0].strip() + "[" + parts[1].strip()
        tag = parts[-1].replace("]", "").strip()
        if tag == "protected":
            protected.append(label)
        elif tag == "non-protected":
            non_protected.append(label)

    result_color = "success" if assessment["fair"] else "danger"

    return f"""
    <div class="card p-3 border-{result_color} border-2 mb-2">
        <div><strong>Protected:</strong> {' AND '.join(protected)}</div>
        <div><strong>Non-Protected:</strong> {' AND '.join(non_protected)}</div>
        <div><strong>Metric:</strong> {metric.metric}</div>
        <div><strong>Threshold:</strong> {metric.threshold}%</div>
        <div><strong>Assessment Score:</strong> {assessment['value']}</div>
        <div><strong>Result:</strong> <span class="text-{result_color}">{'Fair' if assessment['fair'] else 'Unfair'}</span></div>
    </div>
    """, 'Fair' if assessment['fair'] else 'Unfair'


def render_combine_existing_preview(staging):
    designs = DesignCombineMetrics.objects.filter(sid=staging, delete_flag=False).order_by('group_level', 'order')
    if not designs.exists():
        return "<div class='alert alert-warning'>No combined designs found.</div>", "Unfair"

    group_map = defaultdict(list)
    for d in designs:
        group_map[d.group_level or 0].append(d)

    multiple_groups = len(group_map.keys()) > 1
    all_blocks = []
    final_result = False
    resolved_group = None
    group_logic_results = {}
    has_custom_operator = False

    for group_level in sorted(group_map.keys()):
        group_blocks = []
        logic_chain = []

        for d in group_map[group_level]:
            protected = []
            non_protected = []

            try:
                protected_df, non_protected_df, prot_cols, nonprot_cols = parse_feature_tags(
                    d.features,
                    return_columns=True
                )
            except Exception as e:
                block = f"<div class='card p-2 m-1 text-bg-danger'>Error parsing: {str(e)}</div>"
                group_blocks.append(block)
                logic_chain.append((False, d.boolean_operator or "AND"))
                continue

            threshold = (d.threshold or 0.0)
            metric_id = (d.metric or "").lower().strip()

            try:
                if metric_id == "counterfactual_fairness":
                    if not (prot_cols and nonprot_cols):
                        raise ValueError("Counterfactual fairness requires 1 protected and 1 non-protected feature.")
                    col_a, col_b = prot_cols[0], nonprot_cols[0]
                    result = counterfactual_fairness(data, col_a, col_b, threshold=threshold)
                elif metric_id == "conditional_statistical_parity":
                    fn = FAIRNESS_METRIC_MAP[metric_id]
                    result = fn(protected_df, non_protected_df, threshold, group_cols=[
                        "Job", "Credit History", "Savings Account/Bonds"
                    ])
                else:
                    fn = FAIRNESS_METRIC_MAP.get(metric_id, FAIRNESS_METRIC_MAP["demographic_parity"])
                    result = fn(protected_df, non_protected_df, threshold)
            except Exception as e:
                result = {"value": "-", "threshold": threshold, "fair": False}
                block = f"<div class='card p-2 m-1 text-bg-danger'>Metric error: {str(e)}</div>"
                group_blocks.append(block)
                logic_chain.append((False, d.boolean_operator or "AND"))
                continue

            # Extract label text
            for item in d.features.split(","):
                item = item.strip()
                if not item:
                    continue
                parts = item.split("[")
                label = parts[0].strip() + "[" + parts[1].strip()
                tag = parts[-1].replace("]", "").strip()
                if tag == "protected":
                    protected.append(label)
                elif tag == "non-protected":
                    non_protected.append(label)

            block = f"""<div class='card p-2 m-1'>
                <strong>Group Level:</strong> {group_level}<br>
                <strong>Protected:</strong> {' AND '.join(protected)}<br>
                <strong>Non-Protected:</strong> {' AND '.join(non_protected)}<br>
                <strong>Metric:</strong> {d.metric}<br>
                <strong>Threshold:</strong> {d.threshold}%<br>
                <strong>Assessment Score:</strong> {result['value']}<br>
                <strong>Result:</strong> {"Fair" if result['fair'] else "Unfair"}
            </div>"""
            group_blocks.append(block)

            op = d.boolean_operator or "AND"
            if op not in ["AND", "OR"]:
                has_custom_operator = True

            logic_chain.append((result["fair"], op))

        custom_op_found = any(op not in ["AND", "OR"] for _, op in logic_chain)

        if custom_op_found:
            group_result = None
        else:
            group_result = logic_chain[0][0] if logic_chain else False
            prev_op = None
            for val, op in logic_chain:
                if prev_op is None:
                    prev_op = op
                    continue
                if prev_op == "AND":
                    group_result = group_result and val
                elif prev_op == "OR":
                    group_result = group_result or val
                prev_op = op

        group_logic_results[group_level] = group_result

        visual = ""
        for i, block in enumerate(group_blocks):
            visual += block
            if i < len(group_blocks) - 1:
                visual += f"<div class='text-center'><strong>{logic_chain[i][1]}</strong></div>"

        if multiple_groups:
            if group_result is None:
                visual += f"<div class='alert alert-warning'>Group Level {group_level} Result: <strong>Researcher input required due to custom operator</strong></div>"
            else:
                visual += f"<div class='alert alert-secondary'>Group Level {group_level} Result: <strong>{'Fair' if group_result else 'Unfair'}</strong></div>"

        all_blocks.append((group_level, visual))

    # Evaluate final result
    if has_custom_operator:
        final_result = None
    else:
        if multiple_groups:
            for group_level in sorted(group_logic_results.keys()):
                if group_logic_results[group_level]:
                    final_result = True
                    resolved_group = group_level
                    break
            else:
                final_result = False
        else:
            final_result = next(iter(group_logic_results.values()), False)

    final_html = ""
    for _, block_html in all_blocks:
        final_html += block_html + "<hr>"

    result_color = "warning" if final_result is None else "success" if final_result else "danger"
    # Final result rendering
    if final_result is None:
        result_text = "<strong>Final Combined Result:</strong> <span class='text-{result_color}'>Researcher input required due to custom operator.</span>"
    else:
        result_text = f"<strong>Final Combined Result:</strong> <span class='text-{result_color}'>{'Fair' if final_result else 'Unfair'}</span>"
        if multiple_groups:
            result_text += f"<br><strong>Resolved At Group Level:</strong> {resolved_group}"

    
    
    final_html += f"""
        <div  class='alert alert-{result_color}'>
            {result_text}
        </div>
    """

    return final_html, 'Fair' if final_result else 'Unfair'


def render_custom_preview(staging):
    lhs_cards = DesignCustomOwnMetric.objects.filter(sid=staging, side="LHS", delete_flag=False).order_by('order')
    rhs_cards = DesignCustomOwnMetric.objects.filter(sid=staging, side="RHS", delete_flag=False).order_by('order')
    combine_cards = DesignCustomOwnMetric.objects.filter(sid=staging, side__isnull=True, delete_flag=False).order_by('order')

    def build_block(cards, threshold=None):
        blocks = []
        for i, card in enumerate(cards):
            conds = card.conditions.all().order_by('id')
            conditions_str = "<br>".join([
                f"{c.feature} ∈ [{(c.binning or "").strip().split("[")[-1].strip("]")}]" if c.binning else c.feature
                for c in conds
            ])

            block = f"""
            <div class="card mb-2 p-2">
                <strong>Probability Type:</strong> {card.probability_type}<br>
                <strong>Conditions:</strong>
                {conditions_str}<br>
            </div>
            """

            blocks.append(block)

            if card.boolean_operator and i < len(cards) - 1:
                blocks.append(f"<div class='text-center'><strong>{card.boolean_operator}</strong></div>")

        return "\n".join(blocks)

    # Compare Mode: LHS vs RHS
    if lhs_cards.exists() or rhs_cards.exists():
        try:
            global_cfg = DesignCustomOwnGlobal.objects.get(sid=staging)
            threshold = global_cfg.threshold or 0.1
        except DesignCustomOwnGlobal.DoesNotExist:
            threshold = 0.1  # fallback

        result = evaluate_combine_custom_own(lhs_cards, data, threshold)
        result_color = "success" if result["fair"] else "danger"
        result_text = "Fair" if result["fair"] else "Unfair"

        avg_prob_lhs = np.mean([d["probability"] for d in result["details"]]) * 100 if result["details"] else 0.0

        result_html1 = f"""
        <div class='alert alert-{result_color}'>
            <strong>Assessment Score:</strong> <span class='text-{result_color}'>{avg_prob_lhs}</span><br>
            <strong>Threshold:</strong> <span class='text-{result_color}'> {threshold}%</span><br>
            <strong>Final Result:</strong> <span class='text-{result_color}'>{result_text}</span>
        </div>
        """

        result = evaluate_combine_custom_own(rhs_cards, data, threshold)
        result_color = "success" if result["fair"] else "danger"
        result_text = "Fair" if result["fair"] else "Unfair"

        avg_prob_rhs = np.mean([d["probability"] for d in result["details"]]) * 100 if result["details"] else 0.0
        print(f"Avg LHS: {avg_prob_lhs}, Avg RHS: {avg_prob_rhs}")

        result_html2 = f"""
        <div class='alert alert-{result_color}'>
            <strong>Assessment Score:</strong> <span class='text-{result_color}'>{avg_prob_rhs}</span><br>
            <strong>Threshold:</strong> <span class='text-{result_color}'> {threshold}%</span><br>
            <strong>Final Result:</strong> <span class='text-{result_color}'>{result_text}</span>
        </div>
        """

        delta = abs(avg_prob_lhs - avg_prob_rhs)
        vs_fair = delta <= threshold

        vs_color = "success" if vs_fair else "danger"
        vs_text = "Comparable" if vs_fair else "Significantly Different"

        vs_html = f"""
        <div class='alert alert-{vs_color} mt-3'>
            <strong>LHS vs RHS:</strong> {vs_text} (Δ = {round(delta, 2)}%)
        </div>
        """

        return f"""
        <div class='row'>
            <div class='col'>
                <h6 class='text-primary'>LHS</h6>
                {build_block(lhs_cards, threshold)}
                {result_html1}
            </div>
            <div class='col text-center align-self-center'>
                <strong class='text-danger'>VS</strong>
            </div>
            <div class='col'>
                <h6 class='text-primary'>RHS</h6>
                {build_block(rhs_cards, threshold)}
                {result_html2}
            </div>
        </div>
        {vs_html}
        """, 'Fair' if vs_fair else 'Unfair'
    
    # Combine Mode: just render all blocks
    elif combine_cards.exists():
        try:
            global_cfg = DesignCustomOwnGlobal.objects.get(sid=staging)
            threshold = global_cfg.threshold or 0.1
        except DesignCustomOwnGlobal.DoesNotExist:
            threshold = 0.1  # fallback

        result = evaluate_combine_custom_own(combine_cards, data, threshold)
        result_color = "success" if result["fair"] else "danger"
        result_text = "Fair" if result["fair"] else "Unfair"

        result_html = f"""
        <div class='alert alert-{result_color}'>
            <strong>Assessment Score:</strong> <span class='text-{result_color}'>{np.mean([d["probability"] for d in result["details"]]) * 100 if result["details"] else 0.0}</span><br>
            <strong>Threshold:</strong> <span class='text-{result_color}'> {threshold}%</span><br>
            <strong>Final Result:</strong> <span class='text-{result_color}'>{result_text}</span>
        </div>
        """

        return f"""
        <div class='custom-combine-preview'>
            {build_block(combine_cards, threshold)}
            {result_html}
        </div>
        """, result_text
    
    return "No custom logic defined", "Unfair"


def render_procedural_preview(staging, importance_path="feature_importance.csv"):
    try:
        imp_df = pd.read_csv(os.path.join(settings.BASE_DIR, 'superadminpanel', importance_path))
        imp_map = dict(zip(imp_df['feature'], imp_df['importance']))
    except Exception as e:
        return f"<div class='alert alert-danger'>Error loading importance CSV: {e}</div>", "Unfair"

    designs = DesignProceduralMetric.objects.filter(sid=staging, delete_flag=False).order_by('order')
    badges = []

    total = designs.count()
    if total == 0:
        return "<em>No procedural logic defined.</em>", "Unfair"

    weight_per_condition = 100 / total
    cumulative_score = 0.0

    try:
        global_cfg = DesignProceduralGlobal.objects.get(sid=staging)
        threshold = global_cfg.threshold or 60.0
    except DesignProceduralGlobal.DoesNotExist:
        threshold = 60.0

    for d in designs:
        conds = d.conditions.all().order_by('id')
        if not conds.exists():
            continue

        c = conds[0]
        fair = False
        label = ""
        reason = ""

        print(d.label_type)
        if d.label_type == "pre_checks":
            fair = True
            label = f"{c.feature} decision = {c.value}"

        elif d.label_type == "feature_importance":
            actual_imp = imp_map.get(c.feature, 0)
            expected = float(c.value)
            op = c.logic_with_next.strip()
            label = f"{c.feature} importance {op} {expected}"
            if op == ">" and actual_imp > expected:
                fair = True
            elif op == ">=" and actual_imp >= expected:
                fair = True
            elif op == "<" and actual_imp < expected:
                fair = True
            elif op == "<=" and actual_imp <= expected:
                fair = True
            elif op == "==" and actual_imp == expected:
                fair = True
            else:
                fair = False

        elif d.label_type == "does_not_contribute":
            actual_imp = imp_map.get(c.feature, 0)
            fair = actual_imp < 1e-6
            label = f"{c.feature} does NOT contribute"

        elif d.label_type == "contribute":
            actual_imp = imp_map.get(c.feature, 0)
            fair = actual_imp > 0.0
            label = f"{c.feature} contributes"

        elif d.label_type == "importance_by_group":
            base_a = (c.feature or "").strip()
            base_b = (c.value or "").strip()
            op = c.logic_with_next.strip()

            group_a_feats = [f for f in imp_map if f.startswith(f"{base_a}[")]
            group_b_feats = [f for f in imp_map if f.startswith(f"{base_b}[")]

            imp_a_sum = sum([imp_map.get(f, 0) for f in group_a_feats])
            imp_b_sum = sum([imp_map.get(f, 0) for f in group_b_feats])

            label = f"{base_a} group sum {op} {base_b}"

            if op == ">" and imp_a_sum > imp_b_sum:
                fair = True
            elif op == ">=" and imp_a_sum >= imp_b_sum:
                fair = True
            elif op == "<" and imp_a_sum < imp_b_sum:
                fair = True
            elif op == "<=" and imp_a_sum <= imp_b_sum:
                fair = True
            elif op == "==" and imp_a_sum == imp_b_sum:
                fair = True
            else:
                fair = False

        else:
            label = f"{c.feature} ({d.label_type})"
            fair = True

        if fair:
            cumulative_score += weight_per_condition
            badge_color = "success"
        else:
            badge_color = "danger"

        badges.append(f"<span class='badge bg-{badge_color} m-1'>{label}</span>")

    fair_overall = cumulative_score >= threshold
    result_color = "success" if fair_overall else "danger"
    result_text = "Fair" if fair_overall else "Unfair"

    result_html = f"""
    <div class='alert alert-{result_color} mt-3'>
        <strong>Threshold:</strong> {threshold}%<br>
        <strong>Cumulative Score:</strong> {round(cumulative_score, 2)}%<br>
        <strong>Final Result:</strong> <span class='text-{result_color}'>{result_text}</span>
    </div>
    """

    return "<div>" + " ".join(badges) + result_html + "</div>", result_text


def render_affordability_preview(design):
    lhs = design.cards.filter(side="LHS", delete_flag=False).order_by('created_at')
    rhs = design.cards.filter(side="RHS", delete_flag=False).order_by('created_at')

    lhs_mask = parse_affordability_condition_mask(lhs, data, FEATURE_CATEGORY_MAP)
    rhs_mask = parse_affordability_condition_mask(rhs, data, FEATURE_CATEGORY_MAP)

    lhs_score = (data[lhs_mask]["label"] == 1).mean() if lhs_mask.any() else 0.0
    rhs_score = (data[rhs_mask]["label"] == 1).mean() if rhs_mask.any() else 0.0

    lhs_score = round(lhs_score * 100, 2)
    rhs_score = round(rhs_score * 100, 2)

    threshold = design.threshold or 0.0
    diff = round(abs(lhs_score - rhs_score), 2)
    fair = diff <= threshold

    def build_card(side_label, cards):
        conditions = "<br>".join([f"{c.feature} {c.operator} {c.value}" for c in cards]) or "<em>No conditions</em>"
        return f"""
        <div class="card p-3 mb-2">
            <h6 class="text-primary">{side_label}</h6>
            {conditions}
        </div>
        """

    lhs_block = build_card("Credit Conditions (LHS)", lhs)
    rhs_block = build_card("Personal Conditions (RHS)", rhs)

    result_color = "success" if fair else "danger"
    result_text = "Balanced" if fair else "Unbalanced"

    result_html = f"""
        <div class='alert alert-{result_color} mt-2'>
            <strong>Assessment Credit Score:</strong> {lhs_score}%<br>
            <strong>Assessment Personal Score:</strong> {rhs_score}%<br>
            <strong>Assessment Difference:</strong> {diff}%<br>
            <strong>Threshold:</strong> {threshold}%<br>
            <strong>Final Result:</strong> <span class='text-{result_color}'>{result_text}</span>
        </div>
    """

    return f"""
    <div class="row">
        <div class="col-md-5">
            {lhs_block}
        </div>
        <div class="col-md-2 text-center align-self-center">
            <strong class="text-danger">VS</strong>
        </div>
        <div class="col-md-5">
            {rhs_block}
        </div>
        {result_html}
    </div>
    """, 'Fair' if fair else 'Unfair'

def superadmin_dashboard(request):
    staging_data = Staging.objects.values('pid').exclude(pid__isnull=True).exclude(pid__exact='').distinct()
    participants = []

    for entry in staging_data:
        pid = entry['pid']
        sid_set = list(Staging.objects.filter(pid=pid).values_list('id', flat=True))

        existing_count = DesignExistingMetrics.objects.filter(sid__in=sid_set, delete_flag=False).values('sid').distinct().count()
        combine_existing_count = DesignCombineMetrics.objects.filter(sid__in=sid_set, delete_flag=False).values('sid').distinct().count()

        own_combine_count = DesignCustomOwnMetric.objects.filter(
            sid__in=sid_set, side__isnull=True, delete_flag=False
        ).values('sid').distinct().count()

        own_compare_count = DesignCustomOwnMetric.objects.filter(
            sid__in=sid_set, side__in=["LHS", "RHS"], delete_flag=False
        ).values('sid').distinct().count()

        procedural_sids = DesignProceduralMetric.objects.filter(
            sid__in=sid_set, delete_flag=False
        ).values_list('sid', flat=True).distinct()
        procedural_count = Staging.objects.filter(id__in=procedural_sids).count()

        affordability_sids = AffordabilityCard.objects.filter(
            design__staging_id__in=sid_set, delete_flag=False
        ).values_list('design__staging_id', flat=True).distinct()
        affordability_count = Staging.objects.filter(id__in=affordability_sids).count()

        participants.append({
            'pid': pid,
            'existing': existing_count,
            'combine_existing': combine_existing_count,
            'own_combine': own_combine_count,
            'own_compare': own_compare_count,
            'procedural': procedural_count,
            'affordability': affordability_count,
        })

    return render(request, 'superadminpanel/dashboard.html', {'participants': participants})

def superadmin_view_metrics(request, pid):

    staging_objects = []

    if pid:
        staging_objects = Staging.objects.filter(pid=pid).order_by('priority', 'id')

    staging_data = []

    results = list()

    for staging in staging_objects:
        metrics = []

        if staging.metric_type == "existing":
            for m in staging.existing_designs.filter(delete_flag=False):
                preview, result = render_metric_preview(m, staging)
                results.append(result)
                metrics.append({
                    "type": "Existing",
                    "preview": preview,
                    "edit_url": f"{reverse('design_metric_features', args=[staging.id])}?edit={m.id}"
                })

        elif staging.metric_type == "combine_existing":
            preview, result = render_combine_existing_preview(staging)
            results.append(result)
            metrics.append({
                "type": "Combined",
                "preview": preview,
                "edit_url": reverse('design_combine_existing_review_all', args=[staging.id])
            })

        elif staging.metric_type == "custom_own" and staging.perspective == "outcome":
            if(len(DesignCustomOwnMetric.objects.filter(sid=staging, side="LHS", delete_flag=False)) == 0):
                preview, result = render_custom_preview(staging)
                results.append(result)
                metrics.append({
                    "type": "Custom Logic",
                    "preview": preview,
                    "edit_url": reverse('custom_own_logic_review', args=[staging.id])
                })
            else:
                preview, result = render_custom_preview(staging)
                results.append(result)
                metrics.append({
                    "type": "Custom Logic",
                    "preview": preview,
                    "edit_url": reverse('custom_own_compare_final_review', args=[staging.id])
                })

        if staging.perspective == "procedural" and staging.procedural_designs.exists():
            preview, result = render_procedural_preview(staging)
            results.append(result)
            metrics.append({
                "type": "Procedural",
                "preview": preview,
                "edit_url": reverse('procedural_builder', args=[staging.id])
            })

        if hasattr(staging, 'AffordabilityDesign') and staging.AffordabilityDesign and not staging.AffordabilityDesign.delete_flag:
            preview, result = render_affordability_preview(staging.AffordabilityDesign)
            results.append(result)
            metrics.append({
                "type": "Affordability",
                "preview": preview,
                "edit_url": reverse('affordability_builder', args=[staging.id])
            })

        staging_data.append({
            "id": staging.id,
            "category": staging.category,
            "perspective": staging.perspective,
            "metric_type": staging.metric_type,
            "metrics": metrics,
            "priority": staging.priority,
        })

    total_passed = len([r for r in results if r == 'Fair'])
    final_result = "Unfair" if total_passed == 0 else "Fair" if total_passed == len(results) else "Partially Fair ("+str(total_passed)+"/"+str(len(results))+")"
    print(f"Final Result: {final_result}")

    color = "success" if final_result == "Fair" else "danger" if final_result == "Unfair" else "warning"

    final_result_html = f"""
    <div class="alert alert-{color} border border-2">
        <h5 class="mb-0">Overall Fairness Assessment: {final_result}</h5>
    </div>"""

    return render(request, 'superadminpanel/view_metrics.html', {
                            "pid": pid,
                            "staging_data": staging_data,
                            "final_result": final_result_html
                        })

def reorder_staging_view(request, pid):
    stagings = list(Staging.objects.filter(pid=pid).order_by('priority', 'id'))

    if request.method == "POST":
        priorities = {}
        for key, value in request.POST.items():
            if key.startswith("priority_"):
                try:
                    sid = int(key.split("_")[1])
                    prio = int(value)
                    priorities[sid] = prio
                except ValueError:
                    continue

        for staging in stagings:
            staging.priority = priorities.get(staging.id, staging.priority)
            staging.save()

        return HttpResponseRedirect(reverse('superadmin_view_metrics', args=[pid]))

    return render(request, "superadminpanel/reorder_staging.html", {
        "pid": pid,
        "stagings": stagings,
        "total_count": len(stagings)
    })

def export_metrics_excel(request):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "Fairness Metrics Log"

    headers = [
        "Participant ID", "Priority", "Category", "Perspective",
        "Metric Type", "Assessment Result", "Metric Preview"
    ]
    ws.append(headers)

    staging_data = Staging.objects.values('pid').exclude(pid__isnull=True).exclude(pid__exact='').distinct()

    for entry in staging_data:
        pid = entry['pid']

        if("temp" in pid):
            continue
        stagings = Staging.objects.filter(pid=pid).order_by('priority', 'id')

        for staging in stagings:
            metric_rows = []

            if staging.metric_type == "existing":
                for m in staging.existing_designs.filter(delete_flag=False):
                    preview, result = render_metric_preview(m, staging)
                    metric_rows.append(("Existing", result, preview))

            elif staging.metric_type == "combine_existing":
                preview, result = render_combine_existing_preview(staging)
                metric_rows.append(("Combined", result, preview))

            elif staging.metric_type == "custom_own" and staging.perspective == "outcome":
                preview, result = render_custom_preview(staging)
                metric_rows.append(("Custom Logic", result, preview))

            if staging.perspective == "procedural" and staging.procedural_designs.exists():
                preview, result = render_procedural_preview(staging)
                metric_rows.append(("Procedural", result, preview))

            if hasattr(staging, 'AffordabilityDesign') and staging.AffordabilityDesign and not staging.AffordabilityDesign.delete_flag:
                preview, result = render_affordability_preview(staging.AffordabilityDesign)
                metric_rows.append(("Affordability", result, preview))

            for metric_type, result, preview in metric_rows:
                preview_text = strip_tags(preview)
                ws.append([
                    pid,
                    staging.priority,
                    staging.category,
                    staging.perspective,
                    metric_type,
                    result,
                    preview_text
                ])

    # Adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=UI_metrics_export.xlsx'
    wb.save(response)
    return response